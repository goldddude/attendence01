"""
Excel/CSV file parsing utilities
Uses openpyxl for Excel and stdlib csv for CSV — no pandas dependency.
"""
import csv
from openpyxl import load_workbook


REQUIRED_COLUMNS = ['Name', 'Register Number', 'Section', 'Department', 'Duration']


def _normalize_headers(headers):
    """Return a mapping from normalized header -> original index."""
    mapping = {}
    for idx, h in enumerate(headers):
        if h is not None:
            mapping[str(h).strip().lower()] = idx
    return mapping


def _find_column(mapping, col_name):
    """Find a column index by case-insensitive name."""
    return mapping.get(col_name.lower())


def parse_student_file(file_path, file_type='excel'):
    """
    Parse Excel or CSV file containing student data.

    Args:
        file_path: Path to the file
        file_type: 'excel' or 'csv'

    Returns:
        tuple: (success, data_or_error)
            - If success: (True, list of student dictionaries)
            - If error: (False, error_message)
    """
    try:
        rows = []

        if file_type == 'excel':
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            raw_rows = list(ws.values)
            wb.close()
            if not raw_rows:
                return False, "Excel file is empty"
            rows = [[str(cell).strip() if cell is not None else '' for cell in row] for row in raw_rows]
        else:
            with open(file_path, newline='', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                rows = [[cell.strip() for cell in row] for row in reader]

        if not rows:
            return False, "File is empty"

        # First row is the header
        header_row = rows[0]
        mapping = _normalize_headers(header_row)

        # Validate required columns
        missing = []
        col_indices = {}
        for col in REQUIRED_COLUMNS:
            idx = _find_column(mapping, col)
            if idx is None:
                missing.append(col)
            else:
                col_indices[col] = idx

        if missing:
            return False, f"Missing required columns: {', '.join(missing)}"

        # Parse data rows
        students = []
        for row in rows[1:]:
            # Pad row if shorter than expected
            padded = list(row) + [''] * (max(col_indices.values()) + 1 - len(row))

            name = padded[col_indices['Name']]
            reg = padded[col_indices['Register Number']]

            # Skip empty rows
            if not name or not reg or name == 'None' or reg == 'None':
                continue

            student = {
                'name': name,
                'register_number': reg,
                'section': padded[col_indices['Section']],
                'department': padded[col_indices['Department']],
                'duration': padded[col_indices['Duration']],
            }
            students.append(student)

        if not students:
            return False, "No valid student data found in file"

        return True, students

    except Exception as e:
        return False, f"Error parsing file: {str(e)}"


def create_sample_excel(output_path):
    """
    Create a sample Excel template for student data.

    Args:
        output_path: Path where the template should be saved
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    # Headers
    headers = ['Name', 'Register Number', 'Section', 'Department', 'Duration']
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Sample data
    sample_data = [
        ['John Doe', '2021CS001', 'A', 'Computer Science', 'Year 3'],
        ['Jane Smith', '2021CS002', 'A', 'Computer Science', 'Year 3'],
        ['Alice Johnson', '2021EC001', 'B', 'Electronics', 'Year 2'],
    ]

    for row_data in sample_data:
        ws.append(row_data)

    # Adjust column widths
    column_widths = [25, 20, 12, 25, 15]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width

    # Save
    wb.save(output_path)
    print(f"✅ Sample template created: {output_path}")
